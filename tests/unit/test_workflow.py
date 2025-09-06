"""Tests for workflow export/import functionality."""

import json
from pathlib import Path
from datetime import datetime

import pytest

from docdiff.workflow import TranslationExporter, TranslationImporter
from docdiff.compare.models import ComparisonResult, NodeMapping
from docdiff.models import DocumentNode, NodeType, TranslationPair, TranslationStatus


@pytest.fixture
def sample_comparison_result():
    """Create a sample comparison result for testing."""
    source_nodes = [
        DocumentNode(
            id="src_1",
            type=NodeType.SECTION,
            content="# Introduction",
            level=1,
            title="Introduction",
            label="intro",
            file_path=Path("docs/en/index.md"),
            line_number=1,
            content_hash="hash1",
        ),
        DocumentNode(
            id="src_2",
            type=NodeType.PARAGRAPH,
            content="This is a test paragraph.",
            file_path=Path("docs/en/index.md"),
            line_number=3,
            content_hash="hash2",
        ),
        DocumentNode(
            id="src_3",
            type=NodeType.CODE_BLOCK,
            content="print('Hello')",
            language="python",
            name="example",
            file_path=Path("docs/en/index.md"),
            line_number=5,
            content_hash="hash3",
        ),
    ]

    target_nodes = [
        DocumentNode(
            id="tgt_1",
            type=NodeType.SECTION,
            content="# はじめに",
            level=1,
            title="はじめに",
            label="intro",
            file_path=Path("docs/ja/index.md"),
            line_number=1,
            content_hash="hash4",
        ),
        # Paragraph is missing translation
        # Code block is also missing
    ]

    mappings = [
        NodeMapping(
            source_node=source_nodes[0],
            target_node=target_nodes[0],
            similarity=1.0,
            mapping_type="exact",
        ),
        NodeMapping(
            source_node=source_nodes[1],
            target_node=None,
            similarity=0.0,
            mapping_type="missing",
        ),
        NodeMapping(
            source_node=source_nodes[2],
            target_node=None,
            similarity=0.0,
            mapping_type="missing",
        ),
    ]

    translation_pairs = [
        TranslationPair(
            source_node_id="src_1",
            target_node_id="tgt_1",
            source_language="en",
            target_language="ja",
            status=TranslationStatus.TRANSLATED,
        ),
        TranslationPair(
            source_node_id="src_2",
            target_node_id=None,
            source_language="en",
            target_language="ja",
            status=TranslationStatus.PENDING,
        ),
        TranslationPair(
            source_node_id="src_3",
            target_node_id=None,
            source_language="en",
            target_language="ja",
            status=TranslationStatus.PENDING,
        ),
    ]

    return ComparisonResult(
        structure_diff={
            "section": {"source": 1, "target": 1, "diff": 0},
            "paragraph": {"source": 1, "target": 0, "diff": 1},
            "code_block": {"source": 1, "target": 0, "diff": 1},
        },
        content_changes=[],
        translation_pairs=translation_pairs,
        coverage_stats={
            "overall": 0.33,
            "counts": {"total": 3, "translated": 1, "missing": 2, "fuzzy": 0},
        },
        mappings=mappings,
        source_lang="en",
        target_lang="ja",
        timestamp=datetime.now(),
    )


class TestTranslationExporter:
    """Test TranslationExporter functionality."""

    def test_export_csv(self, sample_comparison_result, tmp_path):
        """Test CSV export."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export.csv"

        result_path = exporter.export(
            sample_comparison_result,
            format="csv",
            output_path=output_path,
            options={"include_missing": True},
        )

        assert result_path == output_path
        assert output_path.exists()

        # Load and verify CSV
        import csv

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 2  # Two missing translations

        # Check first missing translation
        first_row = rows[0]
        assert first_row["ID"] == "src_2"
        assert first_row["Status"] == "missing"
        assert first_row["Source"] == "This is a test paragraph."
        assert first_row["Target"] == ""
        assert "File" in first_row
        assert "Line" in first_row
        assert "Notes" in first_row  # Field for translator notes

    def test_export_json(self, sample_comparison_result, tmp_path):
        """Test JSON export."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export.json"

        result_path = exporter.export(
            sample_comparison_result,
            format="json",
            output_path=output_path,
            options={"include_missing": True},
        )

        assert result_path == output_path
        assert output_path.exists()

        # Load and verify JSON
        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert data["metadata"]["source_lang"] == "en"
        assert data["metadata"]["target_lang"] == "ja"
        assert len(data["translations"]) == 2  # Two missing translations

        # Check first missing translation
        trans = data["translations"][0]
        assert trans["id"] == "src_2"
        assert trans["status"] == "missing"
        assert trans["source"] == "This is a test paragraph."
        assert trans["target"] == ""

    def test_export_json_with_context(self, sample_comparison_result, tmp_path):
        """Test JSON export with context information."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export_context.json"

        _ = exporter.export(
            sample_comparison_result,
            format="json",
            output_path=output_path,
            options={"include_all": True, "include_context": True},
        )

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Check context for section node
        section_trans = next(t for t in data["translations"] if t["id"] == "src_1")
        assert section_trans["context"]["label"] == "intro"
        assert section_trans["context"]["title"] == "Introduction"

        # Check context for code block
        code_trans = next(t for t in data["translations"] if t["id"] == "src_3")
        assert code_trans["context"]["name"] == "example"

    def test_export_xlsx(self, sample_comparison_result, tmp_path):
        """Test Excel export."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export.xlsx"

        result_path = exporter.export(
            sample_comparison_result, format="xlsx", output_path=output_path
        )

        assert result_path == output_path
        assert output_path.exists()

        # Verify Excel structure
        import openpyxl

        wb = openpyxl.load_workbook(output_path)

        # Check sheets exist
        assert "Summary" in wb.sheetnames
        assert "Translations" in wb.sheetnames

        # Check summary data
        summary = wb["Summary"]
        assert summary["B4"].value == "en"  # Source language
        assert summary["B5"].value == "ja"  # Target language

        # Check translations
        trans = wb["Translations"]
        # Header + 2 missing translations
        row_count = sum(1 for row in trans.iter_rows(min_row=2) if row[0].value)
        assert row_count == 2

    def test_export_xliff(self, sample_comparison_result, tmp_path):
        """Test XLIFF export."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export.xliff"

        result_path = exporter.export(
            sample_comparison_result, format="xliff", output_path=output_path
        )

        assert result_path == output_path
        assert output_path.exists()

        # Parse and verify XLIFF
        from lxml import etree

        tree = etree.parse(str(output_path))
        root = tree.getroot()

        # Check root attributes
        assert root.get("version") == "2.1"
        assert root.get("srcLang") == "en"
        assert root.get("trgLang") == "ja"

        # Check file elements
        nsmap = {"xliff": "urn:oasis:names:tc:xliff:document:2.1"}
        files = root.xpath("//xliff:file", namespaces=nsmap)
        assert len(files) > 0

        # Check units
        units = root.xpath("//xliff:unit", namespaces=nsmap)
        assert len(units) == 2  # Two missing translations

    def test_unsupported_format(self, sample_comparison_result, tmp_path):
        """Test error on unsupported format."""
        exporter = TranslationExporter()
        output_path = tmp_path / "export.txt"

        with pytest.raises(ValueError, match="Unsupported format"):
            exporter.export(
                sample_comparison_result, format="txt", output_path=output_path
            )


class TestTranslationImporter:
    """Test TranslationImporter functionality."""

    def test_import_csv(self, tmp_path):
        """Test CSV import."""
        import csv

        # Create test CSV file
        csv_path = tmp_path / "import.csv"

        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = [
                "ID",
                "File",
                "Line",
                "Type",
                "Status",
                "Similarity",
                "Source",
                "Target",
                "Label",
                "Name",
                "Notes",
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(
                {
                    "ID": "src_2",
                    "File": "docs/ja/index.md",
                    "Line": "3",
                    "Type": "paragraph",
                    "Status": "translated",
                    "Similarity": "1.00",
                    "Source": "This is a test paragraph.",
                    "Target": "これはテスト段落です。",
                    "Label": "",
                    "Name": "",
                    "Notes": "Translated by test",
                }
            )

        # Import translations
        importer = TranslationImporter()
        target_nodes = []

        updated_nodes, stats = importer.import_translations(
            import_path=csv_path, format="csv", target_nodes=target_nodes
        )

        assert stats["total"] == 1
        assert stats["created"] == 1
        assert len(updated_nodes) == 1
        assert updated_nodes[0].content == "これはテスト段落です。"

    def test_import_json(self, tmp_path):
        """Test JSON import."""
        # Create test JSON file
        test_data = {
            "metadata": {"source_lang": "en", "target_lang": "ja"},
            "translations": [
                {
                    "id": "src_2",
                    "source": "This is a test paragraph.",
                    "target": "これはテスト段落です。",
                    "status": "translated",
                    "type": "paragraph",
                    "file": "docs/ja/index.md",
                    "line": 3,
                }
            ],
        }

        json_path = tmp_path / "import.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(test_data, f)

        # Import translations
        importer = TranslationImporter()
        target_nodes = []

        updated_nodes, stats = importer.import_translations(
            import_path=json_path, format="json", target_nodes=target_nodes
        )

        assert stats["total"] == 1
        assert stats["created"] == 1
        assert len(updated_nodes) == 1
        assert updated_nodes[0].content == "これはテスト段落です。"

    def test_import_xlsx(self, tmp_path):
        """Test Excel import."""
        import openpyxl

        # Create test Excel file
        wb = openpyxl.Workbook()
        trans = wb.active
        trans.title = "Translations"

        # Add headers
        headers = ["ID", "File", "Line", "Type", "Status", "Source", "Target", "Notes"]
        for col, header in enumerate(headers, 1):
            trans.cell(row=1, column=col, value=header)

        # Add translation data
        trans.cell(row=2, column=1, value="src_2")
        trans.cell(row=2, column=2, value="docs/ja/index.md")
        trans.cell(row=2, column=3, value=3)
        trans.cell(row=2, column=4, value="paragraph")
        trans.cell(row=2, column=5, value="translated")
        trans.cell(row=2, column=6, value="Test paragraph")
        trans.cell(row=2, column=7, value="テスト段落")

        xlsx_path = tmp_path / "import.xlsx"
        wb.save(xlsx_path)

        # Import translations
        importer = TranslationImporter()
        target_nodes = []

        updated_nodes, stats = importer.import_translations(
            import_path=xlsx_path, format="xlsx", target_nodes=target_nodes
        )

        assert stats["total"] == 1
        assert stats["created"] == 1
        assert updated_nodes[0].content == "テスト段落"

    def test_import_xliff(self, tmp_path):
        """Test XLIFF import."""
        from lxml import etree

        # Create test XLIFF file
        XLIFF_NS = "urn:oasis:names:tc:xliff:document:2.1"
        nsmap = {None: XLIFF_NS}

        xliff = etree.Element(
            "xliff", version="2.1", srcLang="en", trgLang="ja", nsmap=nsmap
        )

        file_elem = etree.SubElement(
            xliff, "file", id="f1", original="docs/ja/index.md"
        )

        unit = etree.SubElement(file_elem, "unit", id="src_2")
        segment = etree.SubElement(unit, "segment")

        source = etree.SubElement(segment, "source")
        source.text = "Test paragraph"

        target = etree.SubElement(segment, "target")
        target.text = "テスト段落"

        xliff_path = tmp_path / "import.xliff"
        tree = etree.ElementTree(xliff)
        tree.write(str(xliff_path), xml_declaration=True, encoding="utf-8")

        # Import translations
        importer = TranslationImporter()
        target_nodes = []

        updated_nodes, stats = importer.import_translations(
            import_path=xliff_path, format="xliff", target_nodes=target_nodes
        )

        assert stats["total"] == 1
        assert stats["created"] == 1
        assert updated_nodes[0].content == "テスト段落"

    def test_merge_with_comparison(self, sample_comparison_result, tmp_path):
        """Test merging imports with comparison results."""
        # Create import data
        test_data = {
            "translations": [
                {
                    "id": "src_2",
                    "source": "This is a test paragraph.",
                    "target": "これはテスト段落です。",
                    "status": "translated",
                },
                {
                    "id": "src_3",
                    "source": "print('Hello')",
                    "target": "print('こんにちは')",
                    "status": "translated",
                },
            ]
        }

        json_path = tmp_path / "merge.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(test_data, f)

        # Merge with comparison
        importer = TranslationImporter()
        stats = importer.merge_with_comparison(
            import_path=json_path,
            format="json",
            comparison_result=sample_comparison_result,
        )

        assert stats["newly_translated"] == 2
        assert stats["still_missing"] == 0

        # Check mappings are updated
        for mapping in sample_comparison_result.mappings:
            if mapping.source_node.id in ["src_2", "src_3"]:
                assert mapping.target_node is not None
                assert mapping.mapping_type == "translated"

    def test_skip_empty_translations(self, tmp_path):
        """Test skipping empty translations during import."""
        test_data = {
            "translations": [
                {
                    "id": "src_1",
                    "source": "Source text",
                    "target": "",  # Empty translation
                    "status": "missing",
                },
                {
                    "id": "src_2",
                    "source": "Another source",
                    "target": "翻訳済み",
                    "status": "translated",
                },
            ]
        }

        json_path = tmp_path / "skip_empty.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(test_data, f)

        importer = TranslationImporter()
        target_nodes = []

        updated_nodes, stats = importer.import_translations(
            import_path=json_path,
            format="json",
            target_nodes=target_nodes,
            options={"skip_empty": True},
        )

        assert stats["total"] == 1  # Only non-empty translation
        assert stats["created"] == 1
        assert updated_nodes[0].id == "src_2"
