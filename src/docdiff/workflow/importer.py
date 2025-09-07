"""Import functionality for translation workflow."""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from docdiff.compare.models import ComparisonResult, NodeMapping
from docdiff.models import DocumentNode


class TranslationImporter:
    """Import translated content and update documents."""

    def import_translations(
        self,
        import_path: Path,
        format: str,
        target_nodes: List[DocumentNode],
        options: Optional[Dict[str, Any]] = None,
    ) -> Tuple[List[DocumentNode], Dict[str, Any]]:
        """Import translations and apply to target nodes.

        Args:
            import_path: Path to import file
            format: Import format (json, xlsx, xliff)
            target_nodes: Existing target nodes to update
            options: Optional import options

        Returns:
            Tuple of (updated_nodes, import_stats)
        """
        options = options or {}

        if format == "json":
            translations = self._import_json(import_path, options)
        elif format == "csv":
            translations = self._import_csv(import_path, options)
        elif format == "xlsx":
            translations = self._import_xlsx(import_path, options)
        elif format == "xliff":
            translations = self._import_xliff(import_path, options)
        else:
            raise ValueError(f"Unsupported format: {format}")

        # Apply translations to nodes
        updated_nodes, stats = self._apply_translations(
            translations, target_nodes, options
        )

        return updated_nodes, stats

    def _import_json(
        self, import_path: Path, options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Import translations from new hierarchical JSON format.

        Only supports schema version 1.0 format.
        Old flat format is no longer supported.
        """
        with open(import_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Check schema version
        schema_version = data.get("schema_version")
        if not schema_version:
            raise ValueError(
                "Unsupported JSON format. Missing schema_version. "
                "Only schema version 1.0 is supported."
            )

        if schema_version != "1.0":
            raise ValueError(
                f"Unsupported schema version: {schema_version}. "
                "Only schema version 1.0 is supported."
            )

        # Extract translations from either flat or hierarchical structure
        translations = []

        # Check if this is flat format (has translations array) or hierarchical format
        if "translations" in data:
            # Flat format - directly use translations array
            translations = data["translations"]
        else:
            # Hierarchical format - extract from document_hierarchy
            document_hierarchy = data.get("document_hierarchy", {})

            # Process each file
            for file_path, doc_file in document_hierarchy.get("files", {}).items():
                nodes = doc_file.get("nodes", {})

                for node_id, node in nodes.items():
                    # Skip if no translation needed
                    if node.get("status") == "translated" and not options.get(
                        "include_translated", False
                    ):
                        continue

                    # Skip empty translations if specified
                    if options.get("skip_empty") and not node.get("target", "").strip():
                        continue

                    context_data = node.get("context", {})
                    metadata = node.get("metadata", {})

                    translations.append(
                        {
                            "id": node_id,
                            "source": node.get("source", ""),
                            "target": node.get("target", ""),
                            "status": node.get("status", "missing"),
                            "file": context_data.get("file_path", file_path),
                            "line": context_data.get("line_number"),
                            "type": node.get("type"),
                            "context": {
                                "label": metadata.get("label"),
                                "name": metadata.get("name"),
                                "caption": metadata.get("caption"),
                                "parent_section": context_data.get("parent_section"),
                                "preceding_text": context_data.get("preceding_text"),
                                "following_text": context_data.get("following_text"),
                            },
                            "parent_id": node.get("parent_id"),
                            "children_ids": node.get("children_ids", []),
                        }
                    )

        return translations

    def _import_csv(
        self, import_path: Path, options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Import translations from CSV file."""
        translations = []

        with open(import_path, "r", encoding="utf-8", newline="") as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                # Skip empty translations if specified
                if options.get("skip_empty") and not row.get("Target", "").strip():
                    continue

                # Parse line number as integer
                line = row.get("Line", "")
                try:
                    line_number = int(line) if line else None
                except ValueError:
                    line_number = None

                translations.append(
                    {
                        "id": row["ID"],
                        "source": row.get("Source", ""),
                        "target": row.get("Target", ""),
                        "status": row.get("Status", "missing"),
                        "file": row.get("File", ""),
                        "line": line_number,
                        "type": row.get("Type", ""),
                        "context": {
                            "label": row.get("Label", ""),
                            "name": row.get("Name", ""),
                        }
                        if row.get("Label") or row.get("Name")
                        else {},
                    }
                )

        return translations

    def _import_xlsx(
        self, import_path: Path, options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Import translations from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel import. Install with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(import_path)

        # Get translations sheet
        if "Translations" not in wb.sheetnames:
            raise ValueError("Excel file missing 'Translations' sheet")

        trans_sheet = wb["Translations"]
        translations = []

        # Skip header row
        for row in trans_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue

            # Extract data from columns
            # Headers: ID, File, Line, Type, Status, Source, Target, Notes
            target_text = row[6] if row[6] else ""

            # Skip empty translations if specified
            if options.get("skip_empty") and not target_text:
                continue

            translations.append(
                {
                    "id": row[0],
                    "file": row[1] if row[1] else None,
                    "line": row[2] if row[2] else None,
                    "type": row[3] if row[3] else None,
                    "status": row[4] if row[4] else "missing",
                    "source": row[5] if row[5] else "",
                    "target": target_text,
                    "notes": row[7] if len(row) > 7 and row[7] else None,
                }
            )

        return translations

    def _import_xliff(
        self, import_path: Path, options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Import translations from XLIFF file."""
        try:
            from lxml import etree
        except ImportError:
            raise ImportError(
                "lxml is required for XLIFF import. Install with: pip install lxml"
            )

        # Parse XLIFF file
        tree = etree.parse(str(import_path))
        root = tree.getroot()

        # Extract namespace
        nsmap = {"xliff": "urn:oasis:names:tc:xliff:document:2.1"}

        translations = []

        # Process each file element
        for file_elem in root.xpath("//xliff:file", namespaces=nsmap):
            file_path = file_elem.get("original", "unknown")

            # Process each unit
            for unit in file_elem.xpath(".//xliff:unit", namespaces=nsmap):
                unit_id = unit.get("id")

                # Extract notes for context
                context = {}
                notes = unit.xpath(".//xliff:note", namespaces=nsmap)
                for note in notes:
                    if note.text:
                        if note.text.startswith("Label:"):
                            context["label"] = note.text[6:].strip()
                        elif note.text.startswith("Name:"):
                            context["name"] = note.text[5:].strip()
                        elif note.text.startswith("Title:"):
                            context["title"] = note.text[6:].strip()

                # Extract source and target
                segments = unit.xpath(".//xliff:segment", namespaces=nsmap)
                for segment in segments:
                    source_elem = segment.xpath(".//xliff:source", namespaces=nsmap)
                    target_elem = segment.xpath(".//xliff:target", namespaces=nsmap)

                    source_text = (
                        source_elem[0].text
                        if source_elem and source_elem[0].text
                        else ""
                    )
                    target_text = (
                        target_elem[0].text
                        if target_elem and target_elem[0].text
                        else ""
                    )

                    # Skip empty translations if specified
                    if options.get("skip_empty") and not target_text:
                        continue

                    # Determine status from state attribute
                    state = segment.get("state", "initial")
                    if state == "needs-review":
                        status = "fuzzy"
                    elif target_text:
                        status = "translated"
                    else:
                        status = "missing"

                    translations.append(
                        {
                            "id": unit_id,
                            "file": file_path,
                            "source": source_text,
                            "target": target_text,
                            "status": status,
                            "context": context,
                        }
                    )

        return translations

    def _apply_translations(
        self,
        translations: List[Dict[str, Any]],
        target_nodes: List[DocumentNode],
        options: Dict[str, Any],
    ) -> Tuple[List[DocumentNode], Dict[str, Any]]:
        """Apply imported translations to target nodes.

        Args:
            translations: List of translation dictionaries
            target_nodes: Existing target nodes
            options: Import options

        Returns:
            Tuple of (updated_nodes, statistics)
        """
        # Build index of existing nodes
        node_index = {node.id: node for node in target_nodes}

        # Statistics (only count non-empty translations in total)
        non_empty_translations = [
            t for t in translations if t.get("target", "").strip()
        ]
        stats: Dict[str, Any] = {
            "total": len(non_empty_translations),
            "applied": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
        }

        updated_nodes = []
        create_missing = options.get("create_missing", True)
        overwrite_existing = options.get("overwrite_existing", True)

        for trans in translations:
            node_id = trans["id"]

            # Check if node exists
            if node_id in node_index:
                node = node_index[node_id]

                # Update existing node
                if overwrite_existing or not node.content:
                    node.content = trans["target"]
                    node.content_hash = self._calculate_hash(trans["target"])
                    updated_nodes.append(node)
                    stats["updated"] += 1
                    stats["applied"] += 1
                else:
                    stats["skipped"] += 1

            elif create_missing and trans["target"]:
                # Create new node for missing translation
                new_node = self._create_node_from_translation(trans)
                if new_node:
                    updated_nodes.append(new_node)
                    node_index[new_node.id] = new_node
                    stats["created"] += 1
                    stats["applied"] += 1
                else:
                    stats["errors"].append(f"Could not create node for {node_id}")
            else:
                stats["skipped"] += 1

        # Include unchanged nodes
        for node in target_nodes:
            if node.id not in [n.id for n in updated_nodes]:
                updated_nodes.append(node)

        return updated_nodes, stats

    def _create_node_from_translation(
        self, translation: Dict[str, Any]
    ) -> Optional[DocumentNode]:
        """Create a new DocumentNode from translation data."""
        from docdiff.models import NodeType

        # Map string type to NodeType enum
        type_map = {
            "section": NodeType.SECTION,
            "paragraph": NodeType.PARAGRAPH,
            "code_block": NodeType.CODE_BLOCK,
            "list_item": NodeType.LIST_ITEM,
            "list": NodeType.LIST,
            "table": NodeType.TABLE,
            "figure": NodeType.FIGURE,
            "math_block": NodeType.MATH_BLOCK,
            "admonition": NodeType.ADMONITION,
        }

        # Get type safely
        type_str = translation.get("type", "paragraph")
        if type_str:
            type_str = type_str.lower()
        else:
            type_str = "paragraph"

        node_type = type_map.get(type_str, NodeType.PARAGRAPH)

        # Create node with available data
        file_path = translation.get("file")
        if not file_path:
            file_path = "unknown.md"

        node = DocumentNode(
            id=translation["id"],
            type=node_type,
            content=translation["target"],
            content_hash=self._calculate_hash(translation["target"]),
            # Provide file_path and line_number with defaults
            file_path=Path(file_path),
            line_number=translation.get("line", 0)
            if translation.get("line") is not None
            else 0,
            # Required fields with defaults
            level=0,
            title="",
            label=None,
            name=None,
            caption=None,
            language=None,
            parent_id=None,
            source_file_hash="",
            last_modified=None,
        )

        # Add context fields
        context = translation.get("context", {})
        if context.get("label"):
            node.label = context["label"]
        if context.get("name"):
            node.name = context["name"]
        if context.get("title"):
            node.title = context["title"]

        return node

    def _calculate_hash(self, content: str) -> str:
        """Calculate hash for content."""
        import hashlib

        return hashlib.sha256(content.encode("utf-8")).hexdigest()[:16]

    def merge_with_comparison(
        self,
        import_path: Path,
        format: str,
        comparison_result: ComparisonResult,
        options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Merge imported translations with comparison results.

        This method is useful for updating translation status after import.

        Args:
            import_path: Path to import file
            format: Import format
            comparison_result: Previous comparison result
            options: Import options

        Returns:
            Merge statistics
        """
        options = options or {}

        # Import translations
        if format == "json":
            translations = self._import_json(import_path, options)
        elif format == "csv":
            translations = self._import_csv(import_path, options)
        elif format == "xlsx":
            translations = self._import_xlsx(import_path, options)
        elif format == "xliff":
            translations = self._import_xliff(import_path, options)
        else:
            raise ValueError(f"Unsupported format: {format}")

        # Build translation index
        trans_index = {t["id"]: t for t in translations}

        # Update comparison mappings
        stats = {
            "total_mappings": len(comparison_result.mappings),
            "updated": 0,
            "newly_translated": 0,
            "still_missing": 0,
        }

        for mapping in comparison_result.mappings:
            source_id = mapping.source_node.id

            if source_id in trans_index:
                trans = trans_index[source_id]

                if trans["target"]:
                    # Update or create target node
                    if not mapping.target_node:
                        # Create new target node
                        mapping.target_node = self._create_node_from_translation(trans)
                        mapping.mapping_type = "translated"
                        stats["newly_translated"] += 1
                    else:
                        # Update existing target node
                        mapping.target_node.content = trans["target"]
                        mapping.target_node.content_hash = self._calculate_hash(
                            trans["target"]
                        )
                        mapping.mapping_type = "exact"
                        stats["updated"] += 1

                    # Update similarity
                    mapping.similarity = 1.0

            elif mapping.mapping_type == "missing":
                stats["still_missing"] += 1

        # Update coverage statistics
        comparison_result.coverage_stats = self._recalculate_coverage(
            comparison_result.mappings
        )

        return stats

    def _recalculate_coverage(self, mappings: List[NodeMapping]) -> Dict[str, Any]:
        """Recalculate coverage statistics after import."""
        total = len(mappings)
        translated = sum(1 for m in mappings if m.is_translated())
        missing = sum(1 for m in mappings if m.mapping_type == "missing")
        fuzzy = sum(1 for m in mappings if m.mapping_type == "fuzzy")

        return {
            "overall": translated / total if total > 0 else 0,
            "counts": {
                "total": total,
                "translated": translated,
                "missing": missing,
                "fuzzy": fuzzy,
            },
        }
