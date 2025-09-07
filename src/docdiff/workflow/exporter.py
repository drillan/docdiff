"""Export functionality for translation workflow."""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime

from docdiff.compare.models import ComparisonResult, NodeMapping
from docdiff.ai.adaptive_optimizer import AdaptiveBatchOptimizer
from docdiff.ai.glossary import Glossary
from docdiff.export import HierarchyBuilder
from docdiff.models.export_schema import (
    DocumentHierarchy,
    ExportMetadata,
    ExportSchema,
    ExportStatistics,
    GlossaryTermExport,
    CrossReferenceExport,
    SphinxContextExport,
)
from docdiff.sphinx.glossary import GlossaryExtractor
from docdiff.sphinx.references import ReferenceDatabase, ReferenceType
from docdiff.sphinx.project import detect_sphinx_project, export_sphinx_config


class TranslationExporter:
    """Export translation tasks in various formats."""

    def export(
        self,
        result: ComparisonResult,
        format: str,
        output_path: Path,
        options: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """Export comparison results for translation.

        Args:
            result: Comparison results to export
            format: Export format (json, xlsx, xliff)
            output_path: Output file path
            options: Optional export options

        Returns:
            Path to the exported file
        """
        options = options or {}

        if format == "json":
            return self._export_json(result, output_path, options)
        elif format == "csv":
            return self._export_csv(result, output_path, options)
        elif format == "xlsx":
            return self._export_xlsx(result, output_path, options)
        elif format == "xliff":
            return self._export_xliff(result, output_path, options)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_json(
        self, result: ComparisonResult, output_path: Path, options: Dict[str, Any]
    ) -> Path:
        """Export as AI-optimized hierarchical JSON format.

        Complete replacement with new schema v1.0 for AI translation optimization.
        """
        # Build hierarchy from flat mappings
        context_window = options.get("context_window", 3)
        builder = HierarchyBuilder(context_window=context_window)
        hierarchy = builder.build_hierarchy(result.mappings)

        # Extract Sphinx context if available
        sphinx_context = self._extract_sphinx_context_v2(result, options)
        _ = []
        if sphinx_context and sphinx_context.glossary_terms:
            _ = [term.term for term in sphinx_context.glossary_terms]

        # Create optimized batches using adaptive optimizer
        batch_size = options.get("batch_size", 2000)
        source_lang = result.source_lang

        # Always use adaptive optimizer for best efficiency
        adaptive_optimizer = AdaptiveBatchOptimizer(
            target_batch_size=batch_size,
            min_batch_size=max(500, batch_size // 4),
            max_batch_size=batch_size,
            source_lang=source_lang,
            preserve_hierarchy=options.get("preserve_hierarchy", True),
            enable_context=options.get("include_context", False),
            context_window=options.get("context_window", 3),
        )

        # Load glossary if provided
        glossary = None
        glossary_file = options.get("glossary_file")
        if glossary_file and glossary_file.exists():
            glossary = Glossary()
            glossary.load_from_file(glossary_file)

        batches, metrics = adaptive_optimizer.optimize_hierarchy(hierarchy, glossary)

        # Print optimization report if verbose
        if options.get("verbose"):
            print(adaptive_optimizer.get_metrics_report())

        # Create metadata
        metadata = self._create_export_metadata(result, hierarchy)

        # Build final schema
        export_data = ExportSchema(
            schema_version="1.0",
            metadata=metadata,
            sphinx_context=sphinx_context,
            document_hierarchy=hierarchy,
            translation_batches=batches,
        )

        # Write to file
        output_path.parent.mkdir(parents=True, exist_ok=True)
        json_content = export_data.model_dump(
            exclude_none=True, by_alias=False, mode="json"
        )

        # Add backward compatibility for translations field
        translations: List[Dict[str, Any]] = []
        include_context = options.get("include_context", False)
        include_all = options.get("include_all", False)
        include_missing = options.get("include_missing", True)
        include_outdated = options.get("include_outdated", False)

        for mapping in result.mappings:
            # Filter based on options
            if mapping.mapping_type == "missing" and not include_missing:
                continue
            if mapping.mapping_type == "outdated" and not include_outdated:
                continue
            if mapping.mapping_type == "exact" and not include_all:
                continue

            trans_item: Dict[str, Any] = {
                "id": mapping.source_node.id,
                "status": mapping.mapping_type,
                "source": mapping.source_node.content,
                "target": mapping.target_node.content if mapping.target_node else "",
            }

            # Add context if requested
            if include_context:
                context = {}
                if mapping.source_node.label:
                    context["label"] = mapping.source_node.label
                if mapping.source_node.name:
                    context["name"] = mapping.source_node.name
                if hasattr(mapping.source_node, "title") and mapping.source_node.title:
                    context["title"] = mapping.source_node.title
                if context:
                    trans_item["context"] = context

            translations.append(trans_item)

        # Add backward compatibility field
        json_content["translations"] = translations

        # Custom serialization for datetime
        def serialize_datetime(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(
                json_content,
                f,
                ensure_ascii=False,
                indent=2,
                default=serialize_datetime,
            )

        return output_path

    def _extract_sphinx_context(
        self, result: ComparisonResult, options: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Extract Sphinx-specific context from comparison results.

        Args:
            result: Comparison results
            options: Export options

        Returns:
            Sphinx context dictionary or None if not a Sphinx project
        """
        # Detect project root
        source_path = options.get("source_path")
        if not source_path:
            # Try to extract from first node with file_path
            for mapping in result.mappings:
                if hasattr(mapping.source_node, "file_path"):
                    source_path = Path(mapping.source_node.file_path).parent
                    break

        if not source_path:
            return None

        # Detect Sphinx project
        project = detect_sphinx_project(Path(source_path))
        if not project:
            return None

        # Initialize context with project info
        context = export_sphinx_config(project)

        # Extract glossary if requested
        if options.get("extract_glossary", True):
            glossary_extractor = GlossaryExtractor()

            # Process source files for glossary terms
            for doc_file in project.get_source_files()[:50]:  # Limit for performance
                try:
                    content = doc_file.read_text(encoding="utf-8")
                    if doc_file.suffix == ".rst":
                        glossary_extractor.extract_from_rst(content, doc_file)
                    elif doc_file.suffix in [".md", ".myst"]:
                        glossary_extractor.extract_from_myst(content, doc_file)

                    # Also find term references
                    glossary_extractor.find_term_references(content, doc_file)
                except Exception:
                    continue  # Skip files that can't be read

            # Add glossary context
            context["glossary"] = glossary_extractor.export_glossary_context()

        # Extract references if requested
        if options.get("track_references", True):
            ref_db = ReferenceDatabase()

            # Process source files for references
            for doc_file in project.get_source_files()[:50]:  # Limit for performance
                try:
                    content = doc_file.read_text(encoding="utf-8")
                    if doc_file.suffix == ".rst":
                        ref_db.extract_labels_from_rst(content, doc_file)
                        ref_db.extract_references_from_rst(content, doc_file)
                    elif doc_file.suffix in [".md", ".myst"]:
                        ref_db.extract_labels_from_myst(content, doc_file)
                        ref_db.extract_references_from_myst(content, doc_file)
                except Exception:
                    continue  # Skip files that can't be read

            # Add reference context
            context["references"] = ref_db.export_reference_context()

        return context

    def _extract_sphinx_context_v2(
        self, result: ComparisonResult, options: Dict[str, Any]
    ) -> Optional[SphinxContextExport]:
        """Extract Sphinx context for new export format.

        Args:
            result: Comparison results
            options: Export options

        Returns:
            Sphinx context or None if not a Sphinx project
        """
        if not options.get("extract_glossary", True) and not options.get(
            "track_references", True
        ):
            return None

        # Detect project root
        source_path = options.get("source_path")
        if not source_path:
            # Try to extract from first node with file_path
            for mapping in result.mappings:
                if hasattr(mapping.source_node, "file_path"):
                    source_path = Path(mapping.source_node.file_path).parent
                    break

        if not source_path:
            return None

        # Detect Sphinx project
        project = detect_sphinx_project(Path(source_path))
        if not project:
            return None

        sphinx_context = SphinxContextExport(
            project_name=project.config.project or None,
            project_version=project.config.version or None,
            has_myst=project.has_myst,
            has_i18n=project.has_i18n,
        )

        # Extract glossary if requested
        if options.get("extract_glossary", True):
            glossary_extractor = GlossaryExtractor()

            # Process source files for glossary terms
            for doc_file in project.get_source_files()[:50]:  # Limit for performance
                try:
                    content = doc_file.read_text(encoding="utf-8")
                    if doc_file.suffix == ".rst":
                        glossary_extractor.extract_from_rst(content, doc_file)
                    elif doc_file.suffix in [".md", ".myst"]:
                        glossary_extractor.extract_from_myst(content, doc_file)

                    # Also find term references
                    glossary_extractor.find_term_references(content, doc_file)
                except Exception:
                    continue  # Skip files that can't be read

            # Convert to export format
            for term in set(glossary_extractor.terms.values()):
                glossary_export = GlossaryTermExport(
                    term=term.term,
                    definition=term.definition,
                    source_file=str(term.source_file),
                    line_number=term.line_number,
                    aliases=term.aliases,
                    usage_count=sum(
                        1
                        for ref in glossary_extractor.references
                        if ref.term == term.term
                    ),
                )
                sphinx_context.glossary_terms.append(glossary_export)

        # Extract references if requested
        if options.get("track_references", True):
            ref_db = ReferenceDatabase()

            # Process source files for references
            for doc_file in project.get_source_files()[:50]:  # Limit for performance
                try:
                    content = doc_file.read_text(encoding="utf-8")
                    if doc_file.suffix == ".rst":
                        ref_db.extract_labels_from_rst(content, doc_file)
                        ref_db.extract_references_from_rst(content, doc_file)
                    elif doc_file.suffix in [".md", ".myst"]:
                        ref_db.extract_labels_from_myst(content, doc_file)
                        ref_db.extract_references_from_myst(content, doc_file)
                except Exception:
                    continue  # Skip files that can't be read

            # Convert to export format
            for ref in ref_db.references[:100]:  # Limit for performance
                ref_export = CrossReferenceExport(
                    type=ref.ref_type.value
                    if isinstance(ref.ref_type, ReferenceType)
                    else str(ref.ref_type),
                    from_node=ref.node_id,
                    to_label=ref.target,
                    resolved=ref.resolved,
                    source_file=str(ref.source_file),
                    line_number=ref.line_number,
                )
                sphinx_context.cross_references.append(ref_export)

        return sphinx_context

    def _create_export_metadata(
        self, result: ComparisonResult, hierarchy: DocumentHierarchy
    ) -> ExportMetadata:
        """Create export metadata.

        Args:
            result: Comparison results
            hierarchy: Document hierarchy

        Returns:
            Export metadata
        """
        # Calculate statistics
        total_nodes = hierarchy.total_nodes
        missing_nodes = sum(file.missing_nodes for file in hierarchy.files.values())
        outdated_nodes = sum(file.outdated_nodes for file in hierarchy.files.values())
        translated_nodes = total_nodes - missing_nodes - outdated_nodes

        coverage = (translated_nodes / total_nodes * 100) if total_nodes > 0 else 0.0

        statistics = ExportStatistics(
            total_nodes=total_nodes,
            missing=missing_nodes,
            outdated=outdated_nodes,
            translated=translated_nodes,
            total_files=hierarchy.total_files,
            coverage_percentage=coverage,
        )

        return ExportMetadata(
            docdiff_version="0.0.1",
            export_timestamp=datetime.now(),
            source_lang=result.source_lang,
            target_lang=result.target_lang,
            statistics=statistics,
            schema_version="1.0",
        )

    def _export_csv(
        self, result: ComparisonResult, output_path: Path, options: Dict[str, Any]
    ) -> Path:
        """Export as CSV for universal compatibility and simplicity."""
        # Filter options
        include_missing = options.get("include_missing", True)
        include_outdated = options.get("include_outdated", False)
        include_all = options.get("include_all", False)

        # Prepare CSV data
        rows = []

        for mapping in result.mappings:
            # Determine if this mapping should be included
            should_include = (
                include_all
                or (include_missing and mapping.mapping_type == "missing")
                or (
                    include_outdated
                    and mapping.mapping_type == "fuzzy"
                    and mapping.similarity < 0.95
                )
            )

            if should_include:
                row = {
                    "ID": mapping.source_node.id,
                    "File": str(mapping.source_node.file_path)
                    if hasattr(mapping.source_node, "file_path")
                    else "",
                    "Line": mapping.source_node.line_number
                    if hasattr(mapping.source_node, "line_number")
                    else "",
                    "Type": mapping.source_node.type.value
                    if hasattr(mapping.source_node.type, "value")
                    else str(mapping.source_node.type),
                    "Status": mapping.mapping_type,
                    "Similarity": f"{mapping.similarity:.2f}",
                    "Source": mapping.source_node.content,
                    "Target": mapping.target_node.content
                    if mapping.target_node
                    else "",
                    "Label": mapping.source_node.label or "",
                    "Name": mapping.source_node.name or "",
                    "Notes": "",  # Empty field for translator notes
                }
                rows.append(row)

        # Write CSV file
        output_path.parent.mkdir(parents=True, exist_ok=True)

        fieldnames = [
            "ID",
            "File",
            "Line",
            "Type",
            "Status",
            "Similarity",
            "Source",
            "Target",
            "Label",
            "Name",
            "Notes",
        ]

        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        return output_path

    def _export_xlsx(
        self, result: ComparisonResult, output_path: Path, options: Dict[str, Any]
    ) -> Path:
        """Export as Excel for translator-friendly format."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        summary = wb.active
        summary.title = "Summary"

        # Add summary headers
        summary["A1"] = "Translation Summary"
        summary["A1"].font = Font(bold=True, size=14)

        summary["A3"] = "Metric"
        summary["B3"] = "Value"
        summary["A3"].font = Font(bold=True)
        summary["B3"].font = Font(bold=True)

        # Add summary data
        summary["A4"] = "Source Language"
        summary["B4"] = result.source_lang
        summary["A5"] = "Target Language"
        summary["B5"] = result.target_lang
        summary["A6"] = "Overall Coverage"
        overall_coverage = result.coverage_stats["overall"]
        assert isinstance(overall_coverage, (int, float))
        summary["B6"] = f"{overall_coverage:.1%}"
        summary["A7"] = "Total Nodes"
        counts = result.coverage_stats["counts"]
        assert isinstance(counts, dict)
        summary["B7"] = counts["total"]
        summary["A8"] = "Translated"
        summary["B8"] = counts["translated"]
        summary["A9"] = "Missing"
        summary["B9"] = counts["missing"]

        # Translations sheet
        trans_sheet = wb.create_sheet("Translations")

        # Headers
        headers = ["ID", "File", "Line", "Type", "Status", "Source", "Target", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = trans_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Add translation data
        row = 2
        missing_fill = PatternFill(
            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
        )
        outdated_fill = PatternFill(
            start_color="FFF0E6", end_color="FFF0E6", fill_type="solid"
        )

        for mapping in result.mappings:
            if mapping.needs_translation():
                # Determine fill color
                fill = (
                    missing_fill if mapping.mapping_type == "missing" else outdated_fill
                )

                # Add data
                trans_sheet.cell(row=row, column=1, value=mapping.source_node.id)

                if hasattr(mapping.source_node, "file_path"):
                    trans_sheet.cell(
                        row=row, column=2, value=str(mapping.source_node.file_path)
                    )
                    trans_sheet.cell(
                        row=row, column=3, value=mapping.source_node.line_number
                    )

                node_type = (
                    mapping.source_node.type.value
                    if hasattr(mapping.source_node.type, "value")
                    else str(mapping.source_node.type)
                )
                trans_sheet.cell(row=row, column=4, value=node_type)

                status_cell = trans_sheet.cell(
                    row=row, column=5, value=mapping.mapping_type
                )
                status_cell.fill = fill

                trans_sheet.cell(row=row, column=6, value=mapping.source_node.content)
                trans_sheet.cell(
                    row=row,
                    column=7,
                    value=mapping.target_node.content if mapping.target_node else "",
                )
                trans_sheet.cell(
                    row=row, column=8, value=""
                )  # Notes column for translators

                row += 1

        # Adjust column widths
        trans_sheet.column_dimensions["A"].width = 15  # ID
        trans_sheet.column_dimensions["B"].width = 30  # File
        trans_sheet.column_dimensions["C"].width = 10  # Line
        trans_sheet.column_dimensions["D"].width = 15  # Type
        trans_sheet.column_dimensions["E"].width = 15  # Status
        trans_sheet.column_dimensions["F"].width = 50  # Source
        trans_sheet.column_dimensions["G"].width = 50  # Target
        trans_sheet.column_dimensions["H"].width = 30  # Notes

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def _export_xliff(
        self, result: ComparisonResult, output_path: Path, options: Dict[str, Any]
    ) -> Path:
        """Export as XLIFF 2.1 for CAT tool integration."""
        try:
            from lxml import etree
        except ImportError:
            raise ImportError(
                "lxml is required for XLIFF export. Install with: pip install lxml"
            )

        # Create XLIFF namespace
        XLIFF_NS = "urn:oasis:names:tc:xliff:document:2.1"
        nsmap = {None: XLIFF_NS}

        # Create XLIFF root element
        xliff = etree.Element(
            "xliff",
            version="2.1",
            srcLang=result.source_lang,
            trgLang=result.target_lang,
            nsmap=nsmap,
        )

        # Group mappings by file
        files_map: Dict[str, List[NodeMapping]] = {}
        for mapping in result.mappings:
            if not mapping.needs_translation():
                continue

            file_path = "unknown"
            if hasattr(mapping.source_node, "file_path"):
                file_path = str(mapping.source_node.file_path)

            if file_path not in files_map:
                files_map[file_path] = []
            files_map[file_path].append(mapping)

        # Create file elements
        for file_idx, (file_path, mappings) in enumerate(files_map.items(), 1):
            file_elem = etree.SubElement(
                xliff, "file", id=f"f{file_idx}", original=file_path
            )

            # Add units for each mapping
            for mapping in mappings:
                unit = etree.SubElement(file_elem, "unit", id=mapping.source_node.id)

                # Add notes if there's metadata
                if mapping.source_node.label or mapping.source_node.name:
                    notes = etree.SubElement(unit, "notes")

                    if mapping.source_node.label:
                        note = etree.SubElement(notes, "note")
                        note.text = f"Label: {mapping.source_node.label}"

                    if mapping.source_node.name:
                        note = etree.SubElement(notes, "note")
                        note.text = f"Name: {mapping.source_node.name}"

                    if (
                        hasattr(mapping.source_node, "title")
                        and mapping.source_node.title
                    ):
                        note = etree.SubElement(notes, "note")
                        note.text = f"Title: {mapping.source_node.title}"

                # Create segment
                segment = etree.SubElement(unit, "segment")

                # Add source
                source = etree.SubElement(segment, "source")
                source.text = mapping.source_node.content

                # Add target if exists (for fuzzy matches)
                if mapping.target_node:
                    target = etree.SubElement(segment, "target")
                    target.text = mapping.target_node.content

                    # Add state attribute for fuzzy matches
                    if mapping.mapping_type == "fuzzy":
                        segment.set("state", "needs-review")

        # Write XLIFF file
        output_path.parent.mkdir(parents=True, exist_ok=True)
        tree = etree.ElementTree(xliff)
        tree.write(
            str(output_path), pretty_print=True, xml_declaration=True, encoding="utf-8"
        )

        return output_path
